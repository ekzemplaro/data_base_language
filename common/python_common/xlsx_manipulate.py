# -*- coding: utf-8 -*-
#
#	xlsx_manipulate.py
#
#						May/31/2017
#
# -------------------------------------------------------------------
import sys
import os

from openpyxl import Workbook
from openpyxl import load_workbook
# -------------------------------------------------------------------
def xlsx_read_proc (xlsx_file):
	dict_aa = {}
	wb = load_workbook (filename = xlsx_file)
	ws = wb.active

	print (ws.max_row)
	max_row = ws.max_row
	for row in ws.rows:
		unit_aa = {}
		key = row[0].value
		unit_aa['name'] = row[1].value
		unit_aa['population'] = int (row[2].value)
		unit_aa['date_mod'] = row[3].value
		dict_aa[key] = unit_aa
#
	return	dict_aa
# -------------------------------------------------------------------
def xlsx_write_proc (xlsx_file,dict_aa):
	wb = Workbook()
	ws = wb.worksheets[0]
	ws.title = "Cities"
	print ("ws.max_row = %d" % ws.max_row)
#
	keys_list = list(dict_aa.keys())
	print ("len = %d\n" % len(keys_list)) 
#
	for it in range (len(keys_list)):
		jt = it + 1
		key = keys_list[it]
		unit_aa = dict_aa[key]
#
		ws['A%d' % jt] = key
		ws['B%d' % jt] = unit_aa['name']
		ws['C%d' % jt] = unit_aa['population']
		ws['D%d' % jt] = unit_aa['date_mod']
#
	wb.save (filename = xlsx_file)
#
	os.chmod (xlsx_file,0o766)
# -------------------------------------------------------------------
